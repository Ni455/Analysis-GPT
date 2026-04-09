# core/morning_scalp.py
import sys, io, os, time, math, logging
import requests
import pytz
import openpyxl
from scipy.stats import norm
from scipy.optimize import brentq

from datetime import datetime, date
from dotenv import load_dotenv

load_dotenv()

from config import (
    GAP_THRESHOLD, TREND_THRESHOLD,
    BUDGET_MIN, BUDGET_MAX, RISK_FREE,
    OTM_IDEAL_MIN, OTM_IDEAL_MAX, OTM_GOOD_MAX, OTM_FAR_MAX,
    HIGH_OI_PCT, LOW_IV_THRESH,
    MORNING_WINDOW_START, MORNING_WINDOW_END,
    EOD_WINDOW_START, EOD_WINDOW_END,
    RATE_LIMIT_SLEEP,
    MIN_SCORE,
)

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)


def compute_direction(gap_pct, trend_pct):
    """
    Returns dict: { signal: "PE"|"CE"|"BOTH", strong: bool, gap_available: bool }
    gap_pct=None means yfinance open data was stale.
    """
    gap_available = gap_pct is not None

    gap_up   = gap_available and gap_pct >= GAP_THRESHOLD
    gap_down = gap_available and gap_pct <= -GAP_THRESHOLD
    trend_up   = trend_pct >= TREND_THRESHOLD
    trend_down = trend_pct <= -TREND_THRESHOLD

    bullish = gap_up or trend_up
    bearish = gap_down or trend_down

    if bullish and bearish:
        return {"signal": "BOTH", "strong": False, "gap_available": gap_available}
    elif bullish:
        strong = gap_up and trend_up
        return {"signal": "PE", "strong": strong, "gap_available": gap_available}
    elif bearish:
        strong = gap_down and trend_down
        return {"signal": "CE", "strong": strong, "gap_available": gap_available}
    else:
        return {"signal": "BOTH", "strong": False, "gap_available": gap_available}


def score_candidate(c):
    """
    Score a single option candidate dict.
    c keys: type, oi, chain_ois, iv, dist_pct, direction (dict)
    Returns int score 0–13.
    """
    score = 0
    sig   = c["direction"]["signal"]
    strong = c["direction"]["strong"]
    opt   = c["type"]   # "CE" or "PE"

    # 1. Trend/gap alignment
    aligned = (sig == opt) or (sig == "BOTH")
    if aligned:
        if strong:
            score += 3      # strongly aligned (replaces weak +1)
            score += 2      # strong bonus
        else:
            score += 1      # weakly aligned

    # 2. OTM distance
    d = c["dist_pct"]
    if OTM_IDEAL_MIN <= d <= OTM_IDEAL_MAX:
        score += 3
    elif OTM_IDEAL_MAX < d <= OTM_GOOD_MAX:
        score += 2
    elif d < OTM_IDEAL_MIN or (OTM_GOOD_MAX < d <= OTM_FAR_MAX):
        score += 1
    # beyond FAR_MAX → 0 (but these are filtered out upstream)

    # 3. High OI (top 30% of chain) — NOT gated by alignment
    chain_ois = c.get("chain_ois", [c["oi"]])
    threshold = sorted(chain_ois)[int(len(chain_ois) * HIGH_OI_PCT)]
    if c["oi"] >= threshold:
        score += 2

    # 4. Low IV — NOT gated by alignment
    if c.get("iv") and c["iv"] < LOW_IV_THRESH:
        score += 1

    return score


def pick_expiry(expiries, today, mode):
    """
    expiries: list of 'YYYY-MM-DD' strings (sorted ascending)
    mode: 'morning' | 'eod'
    Returns expiry string or None if no valid expiry found.
    """
    future = [e for e in sorted(expiries)
              if date.fromisoformat(e) > today]
    if not future:
        return None

    if mode == "morning":
        return future[0]
    else:   # eod → next-after-nearest
        return future[1] if len(future) >= 2 else future[0]


def detect_run_mode(current_time=None):
    """
    Returns 'morning', 'eod', or None if outside valid windows.
    current_time: datetime.time object (defaults to now IST).
    Uses MORNING_WINDOW_START/END and EOD_WINDOW_START/END from config.
    """
    from datetime import time as dtime
    if current_time is None:
        ist = pytz.timezone("Asia/Kolkata")
        current_time = datetime.now(ist).time()

    morning_start = dtime(*MORNING_WINDOW_START)
    morning_end   = dtime(*MORNING_WINDOW_END)
    eod_start     = dtime(*EOD_WINDOW_START)
    eod_end       = dtime(*EOD_WINDOW_END)

    if morning_start <= current_time < morning_end:
        return "morning"
    elif eod_start <= current_time < eod_end:
        return "eod"
    return None


def build_subject(mode, count, run_date=None):
    """
    Builds email subject line.
    mode: 'morning' | 'eod'
    count: number of candidates
    run_date: string like '03-Apr-2026' (defaults to today)
    """
    if run_date is None:
        run_date = date.today().strftime("%d-%b-%Y")
    time_str = "9:27 AM" if mode == "morning" else "3:04 PM"
    label    = "Morning Scalp" if mode == "morning" else "EOD Scalp"
    return f"\U0001f514 {label} Report \u2014 {run_date} | {time_str} | {count} candidates"


def build_email_body(candidates, skipped, mode):
    """
    Builds report body (plain text).
    candidates: list of candidate dicts with symbol, strike, type, ltp, lot_cost, etc.
    skipped: list of dicts with symbol and reason
    mode: 'morning' | 'eod'
    Returns plain text string.
    """
    lines = []
    W = 78
    SEP = "=" * W
    label = "MORNING SCALP \u2014 9:27 AM" if mode == "morning" else "EOD SCALP \u2014 3:04 PM"

    # ── Part 1: Quick action table ─────────────────────────────────────
    lines.append(SEP)
    lines.append(f"  {label}  |  Budget: \u20b910,000\u2013\u20b914,000/lot")
    lines.append(SEP)

    if not candidates:
        lines.append("  No candidates found for today's run.")
    else:
        hdr = f"  {'#':<3} {'Symbol':<12} {'Strike':<8} {'T':<3} {'Score':>6}  {'LTP':>6}  {'LotCost':>9}  {'Gap%':>6}  Signal"
        lines.append(hdr)
        lines.append("  " + "\u2500" * (W - 2))
        for i, c in enumerate(candidates, 1):
            gap_str = f"{c['gap_pct']:+.1f}%" if c.get("gap_pct") is not None else "  N/A"
            arrow   = "\u2191" if c["type"] == "PE" else "\u2193"
            sig_str = f"{arrow} {c['direction']['signal']}{' STRONG' if c['direction']['strong'] else ''}"
            stars   = "\u2605" * c["score"]
            lines.append(
                f"  {i:<3} {c['symbol']:<12} {int(c['strike']):<8} {c['type']:<3} "
                f"{stars:>6}  {c['ltp']:>6.2f}  {c['lot_cost']:>9,.0f}  {gap_str:>6}  {sig_str}"
            )

    lines.append(SEP)
    lines.append("  \u26a0 Confirm LTP on Sensibull/Zerodha before placing order")
    lines.append("")

    # ── Skipped stocks ─────────────────────────────────────────────────
    if skipped:
        lines.append("  SKIPPED / NO CANDIDATES:")
        for s in skipped:
            lines.append(f"    {s['symbol']:<14} \u2014 {s['reason']}")
        lines.append("")

    # ── Part 2: Detail per candidate ───────────────────────────────────
    if candidates:
        lines.append(SEP)
        lines.append("  DETAILED ANALYSIS")
        lines.append(SEP)
        for c in candidates:
            gap_str   = f"{c['gap_pct']:+.1f}%" if c.get("gap_pct") is not None else "N/A"
            trend_str = f"{c['trend_pct']:+.1f}%" if c.get("trend_pct") is not None else "N/A"
            iv_str    = f"{c['iv']:.1f}%" if c.get("iv") is not None else "N/A"
            lines.append(f"\n  {c['symbol']}  {int(c['strike'])} {c['type']}  |  Expiry: {c['expiry']}")
            lines.append(f"  {'─'*55}")
            lines.append(f"  Spot:    \u20b9{c['spot']:,.2f}  |  Gap: {gap_str}  |  5d trend: {trend_str}")
            lines.append(f"  LTP:     \u20b9{c['ltp']:.2f}  |  Lot: {c['lot_size']:,}  |  Cost: \u20b9{c['lot_cost']:,.0f}")
            lines.append(f"  OI:      {c['oi']:,}  |  Vol: {c['volume']:,}  |  IV: {iv_str}")
            lines.append(f"  Score:   {'★' * c['score']}  ({c['score']} pts)")
            lines.append(f"  Signals:")
            for r in c.get("reasons", []):
                lines.append(f"    \u2022 {r}")

    lines.append(f"\n{'=' * W}")
    lines.append("  \u26a0 Data from yfinance + Sensibull. Not financial advice.")
    lines.append(f"{'=' * W}")

    return "\n".join(lines)


def run_scalp(symbols: list[str], mode: str = "morning") -> str:
    """
    Scan the given symbols and return the full report as a plain-text string.
    mode: 'morning' | 'eod'
    """
    import time as _time
    all_candidates = []
    all_skipped = []

    for sym in symbols:
        cands, skip = scan_symbol(sym, mode)
        if cands:
            all_candidates.extend(cands)
        if skip:
            all_skipped.append(skip)
        _time.sleep(RATE_LIMIT_SLEEP)

    all_candidates.sort(key=lambda c: (
        c["score"],
        c.get("oi", 0),
        -(c.get("iv") or 99),
        -(c.get("dist_pct") or 99),
    ), reverse=True)

    top_candidates = [c for c in all_candidates if c["score"] >= MIN_SCORE]
    if not top_candidates:
        top_candidates = all_candidates[:5]

    return build_email_body(top_candidates, all_skipped, mode)


def is_nse_holiday(check_date=None):
    """
    Returns True if check_date is an NSE trading holiday.
    Uses NSE's public holiday API. Falls back to False on error (run anyway).
    """
    if check_date is None:
        check_date = date.today()
    try:
        r = requests.get(
            "https://www.nseindia.com/api/holiday-master?type=trading",
            headers={
                "User-Agent": "Mozilla/5.0",
                "Accept": "application/json",
                "Referer": "https://www.nseindia.com/",
            },
            timeout=10,
        )
        data = r.json()
        # Response: { "CM": [ { "tradingDate": "17-Mar-2026", ... }, ... ] }
        holidays = data.get("CM", [])
        for h in holidays:
            td = h.get("tradingDate", "")
            try:
                hd = datetime.strptime(td, "%d-%b-%Y").date()
                if hd == check_date:
                    return True
            except ValueError:
                continue
        return False
    except Exception as e:
        log.warning(f"NSE holiday check failed ({e}) — assuming trading day")
        return False


def get_stock_data(symbol):
    """
    Returns dict: { spot, prev_close, gap_pct, trend_pct } or None on error.
    gap_pct is None if open data is stale (open == prev_close or zero).
    """
    import yfinance as yf
    try:
        t    = yf.Ticker(f"{symbol}.NS")
        hist = t.history(period="15d")
        if hist.empty or len(hist) < 6:
            log.warning(f"{symbol}: insufficient history")
            return None

        spot       = float(hist["Close"].iloc[-1])
        prev_close = float(hist["Close"].iloc[-2])
        open_price = float(hist["Open"].iloc[-1])

        # gap% — stale if open equals prev_close (yfinance not updated yet)
        if open_price == 0 or abs(open_price - prev_close) < 0.01:
            log.warning(f"[WARNING] Gap data unavailable for {symbol} — using trend only")
            gap_pct = None
        else:
            gap_pct = (open_price - prev_close) / prev_close * 100

        # 5d trend — use yesterday as endpoint to exclude today's incomplete candle
        close_5d_ago    = float(hist["Close"].iloc[-7])   # 5 trading days before yesterday
        yesterday_close = prev_close
        trend_pct = (yesterday_close - close_5d_ago) / close_5d_ago * 100

        return {
            "spot": spot,
            "prev_close": prev_close,
            "gap_pct": gap_pct,
            "trend_pct": trend_pct,
        }
    except Exception as e:
        log.warning(f"{symbol}: yfinance error — {e}")
        return None


def _bs_price(S, K, T, r, sigma, opt):
    d1 = (math.log(S / K) + (r + 0.5 * sigma**2) * T) / (sigma * math.sqrt(T))
    d2 = d1 - sigma * math.sqrt(T)
    if opt == "CE":
        return S * norm.cdf(d1) - K * math.exp(-r * T) * norm.cdf(d2)
    return K * math.exp(-r * T) * norm.cdf(-d2) - S * norm.cdf(-d1)

def calc_iv(S, K, T, r, price, opt):
    """Compute implied volatility via Black-Scholes. Returns % or None."""
    if price <= 0.05 or T <= 0:
        return None
    try:
        return round(brentq(lambda v: _bs_price(S, K, T, r, v, opt) - price, 0.001, 10.0) * 100, 2)
    except Exception:
        return None


def scan_symbol(symbol, mode):
    """
    Full per-symbol scan. Returns (list_of_candidates, skip_dict_or_None).
    mode: 'morning' | 'eod'
    """
    stock_data = get_stock_data(symbol)
    if stock_data is None:
        return [], {"symbol": symbol, "reason": "yfinance error"}

    chain_raw = fetch_options_chain(symbol)
    if not chain_raw:
        return [], {"symbol": symbol, "reason": "API error (timeout)"}

    expiries = sorted(set(i["expiry"] for i in chain_raw))
    expiry   = pick_expiry(expiries, date.today(), mode)
    if not expiry:
        return [], {"symbol": symbol, "reason": "no valid future expiry"}

    expiry_dt = date.fromisoformat(expiry)
    T = max((expiry_dt - date.today()).days, 1) / 365.0

    data   = [i for i in chain_raw if i["expiry"] == expiry]
    ce_map = {i["strike"]: i for i in data if i["instrument_type"] == "CE"}
    pe_map = {i["strike"]: i for i in data if i["instrument_type"] == "PE"}
    strikes = sorted(set(ce_map) | set(pe_map))
    if not strikes:
        return [], {"symbol": symbol, "reason": "no strikes found"}

    # Lot size from API (authoritative — not from Excel)
    sample   = (list(ce_map.values()) + list(pe_map.values()))[0]
    lot_size = int(sample.get("lot_size", 0))
    if not lot_size:
        return [], {"symbol": symbol, "reason": "lot size unavailable"}

    # Budget filter thresholds
    min_ltp = BUDGET_MIN / lot_size
    max_ltp = BUDGET_MAX / lot_size

    spot      = stock_data["spot"]
    gap_pct   = stock_data["gap_pct"]
    trend_pct = stock_data["trend_pct"]
    direction = compute_direction(gap_pct, trend_pct)

    # All OI values for percentile scoring
    all_oi = [int(i.get("oi", 0)) for i in data]

    atm     = min(strikes, key=lambda k: abs(k - spot))
    atm_idx = strikes.index(atm)
    window  = strikes[max(0, atm_idx - 12): atm_idx + 13]

    # PCR for context
    total_ce_oi = sum(int(ce_map[k].get("oi", 0)) for k in ce_map)
    total_pe_oi = sum(int(pe_map[k].get("oi", 0)) for k in pe_map)
    pcr = round(total_pe_oi / total_ce_oi, 3) if total_ce_oi else 1.0

    candidates = []

    for k in window:
        for opt_type in ("PE", "CE"):
            opt_map = pe_map if opt_type == "PE" else ce_map
            opt = opt_map.get(k)
            if not opt:
                continue

            ltp = float(opt.get("last_price", 0) or 0)
            if ltp < min_ltp or ltp > max_ltp:
                continue   # budget filter

            oi  = int(opt.get("oi", 0))
            vol = int(opt.get("volume", 0))
            iv  = calc_iv(spot, k, T, RISK_FREE, ltp, opt_type)

            dist_pct = (
                (spot - k) / spot * 100 if opt_type == "PE"
                else (k - spot) / spot * 100
            )

            if dist_pct < 0:
                continue   # ITM option — skip entirely

            if dist_pct > OTM_FAR_MAX:
                continue   # too far OTM

            c = {
                "symbol": symbol, "expiry": expiry,
                "strike": k, "type": opt_type,
                "ltp": ltp, "lot_size": lot_size,
                "lot_cost": round(ltp * lot_size, 0),
                "oi": oi, "volume": vol, "iv": iv,
                "spot": spot, "gap_pct": gap_pct, "trend_pct": trend_pct,
                "dist_pct": dist_pct, "pcr": pcr,
                "chain_ois": all_oi,
                "direction": direction,
            }
            c["score"]   = score_candidate(c)
            c["reasons"] = _build_reasons(c)
            candidates.append(c)

    if not candidates:
        return [], {"symbol": symbol, "reason": "no strikes in budget range"}

    return candidates, None


def _build_reasons(c):
    """Build human-readable reason strings for a candidate."""
    reasons = []
    sig = c["direction"]["signal"]
    strong = c["direction"]["strong"]
    gap_available = c["direction"]["gap_available"]
    opt = c["type"]

    if not gap_available:
        reasons.append("Gap data unavailable — using trend only")

    if sig == opt or sig == "BOTH":
        if strong:
            g = f"Gap {c['gap_pct']:+.1f}%" if gap_available and c.get("gap_pct") is not None else ""
            t = f"5d trend {c['trend_pct']:+.1f}%"
            reasons.append(f"STRONG signal: {g}{' + ' if g else ''}{t} → {opt} aligned")
        elif sig == opt:
            if gap_available and c.get("gap_pct") is not None and (
                (opt == "PE" and c["gap_pct"] >= 1.0) or (opt == "CE" and c["gap_pct"] <= -1.0)
            ):
                reasons.append(f"Gap {c['gap_pct']:+.1f}% → {opt} signal")
            else:
                reasons.append(f"5d trend {c['trend_pct']:+.1f}% → {opt} signal")

    d = c["dist_pct"]
    if OTM_IDEAL_MIN <= d <= OTM_IDEAL_MAX:
        reasons.append(f"Ideal OTM distance: {d:.1f}% from spot")
    elif OTM_IDEAL_MAX < d <= OTM_GOOD_MAX:
        reasons.append(f"Good OTM distance: {d:.1f}%")
    else:
        reasons.append(f"OTM distance: {d:.1f}%")

    if c.get("iv") is not None and c["iv"] < LOW_IV_THRESH:
        reasons.append(f"Low IV: {c['iv']}% → relatively cheap premium")

    return reasons


_session = requests.Session()
_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept": "application/json",
    "Referer": "https://web.sensibull.com/",
}


def fetch_options_chain(symbol):
    """
    Returns list of instrument dicts from Sensibull, or [] on error.
    Each dict has: strike, instrument_type, expiry, last_price, oi, volume, lot_size
    """
    try:
        r = _session.get(
            f"https://api.sensibull.com/v1/instruments/{symbol}",
            headers=_HEADERS, timeout=15,
        )
        return r.json().get("data", [])
    except Exception as e:
        log.warning(f"{symbol}: Sensibull fetch error — {e}")
        return []


def read_stock_list(xlsx_path):
    """
    Reads stocks.xlsx. Expects column 'Symbol' (case-insensitive).
    Returns list of symbol strings. Extra columns are ignored.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

    PRIORITY_ALIASES = ["symbol", "nse symbol", "nse_symbol", "nsesymbol", "ticker", "scrip"]
    FALLBACK_ALIASES = ["name", "stock"]
    sym_col = next((i for i, h in enumerate(headers) if h in PRIORITY_ALIASES), None)
    if sym_col is None:
        sym_col = next((i for i, h in enumerate(headers) if h in FALLBACK_ALIASES), None)
    if sym_col is None:
        raise ValueError(f"stocks.xlsx must have a 'Symbol' (or 'NSE Symbol') column. Found: {[h for h in headers if h]}")
    symbols = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[sym_col]
        if val:
            symbols.append(str(val).strip().upper())
    wb.close()
    return [s for s in symbols if s]


XLSX_PATH = os.path.join(os.path.dirname(__file__), "stocks.xlsx")

# Fallback symbols used when stocks.xlsx is absent (e.g. cloud/remote runs)
DEFAULT_SYMBOLS = [
    "ITC","ONGC","KOTAKBANK","NTPC","ADANIPOWER","BEL","COALINDIA","POWERGRID",
    "TATASTEEL","ETERNAL","HINDZINC","WIPRO","IOC","JIOFIN","VBL","PFC",
    "UNIONBANK","DLF","BANKBARODA","TATAPOWER","BPCL","PNB","IRFC","CANBK",
    "MOTHERSON","INDUSTOWER","TATAMOTORSDVR","AMBUJACEM","GMRAIRPORT","GAIL",
    "IDEA","ASHOKLEY","BHEL","JSWENERGY","RECLTD","ABCAPITAL","OIL","SWIGGY",
    "NHPC","DABUR","NATIONALUM","ICICIPRULI","NYKAA","HINDPETRO","NMDC",
    "FEDERALBNK","SAIL","BANKINDIA","LTFH","BIOCON","YESBANK","SUZLON","RVNL",
    "IDFCFIRSTB","PATANJALI","VISHAL","KALYANKJIL","PETRONET","HUDCO","CONCOR",
    "IREDA","DELHIVERY","SONACOMS","JUBLFOOD","LICHSGFIN","EXIDEIND","BANDHANBNK",
    "NBCC","TATATECH","ANGELONE","MANAPPURAM","PIRPHARMA","RBLBANK","CROMPTON",
    "INOXWIND","PGEL","SAMMAAN","IEX",
]


def main():
    ist = pytz.timezone("Asia/Kolkata")
    now_ist = datetime.now(ist)

    # 1. Run mode detection
    mode = detect_run_mode(now_ist.time())
    if mode is None:
        log.error(f"Run time outside expected window ({now_ist.strftime('%H:%M')} IST) — exiting")
        sys.exit(1)
    log.info(f"Run mode: {mode.upper()}  |  {now_ist.strftime('%H:%M IST')}")

    # 2. Holiday check
    today = now_ist.date()
    if is_nse_holiday(today):
        log.info(f"NSE Holiday on {today} — skipping")
        sys.exit(0)

    # 3. Read stock list (fall back to DEFAULT_SYMBOLS if xlsx missing)
    if os.path.exists(XLSX_PATH):
        try:
            symbols = read_stock_list(XLSX_PATH)
        except Exception as e:
            log.error(f"Failed to read stocks.xlsx: {e}")
            sys.exit(1)
    else:
        log.warning("stocks.xlsx not found — using built-in DEFAULT_SYMBOLS list")
        symbols = DEFAULT_SYMBOLS
    log.info(f"Scanning {len(symbols)} symbols...")

    # 4. Per-symbol scan
    all_candidates = []
    all_skipped    = []

    for sym in symbols:
        log.info(f"  {sym}...")
        cands, skip = scan_symbol(sym, mode)
        if cands:
            all_candidates.extend(cands)
        if skip:
            all_skipped.append(skip)
        time.sleep(RATE_LIMIT_SLEEP)

    # 5. Sort: score desc, then OI desc, then IV asc (lower = cheaper), then dist_pct asc (closer to ATM)
    all_candidates.sort(key=lambda c: (
        c["score"],
        c.get("oi", 0),
        -(c.get("iv") or 99),
        -(c.get("dist_pct") or 99),
    ), reverse=True)

    # 5b. Keep only high-probability candidates (score >= MIN_SCORE)
    top_candidates = [c for c in all_candidates if c["score"] >= MIN_SCORE]
    if not top_candidates:
        top_candidates = all_candidates[:5]   # fallback: send top 5 if nothing qualifies
    log.info(f"Filtered to {len(top_candidates)} high-probability candidates (min score {MIN_SCORE})")

    # 6. Build report and send via email (legacy path — Telegram bot uses run_scalp() instead)
    run_date = now_ist.strftime("%d-%b-%Y")
    subject  = build_subject(mode, len(top_candidates), run_date)
    body     = build_email_body(top_candidates, all_skipped, mode)

    log.info(f"Report built: {subject}")
    log.info(body)
    log.info("Done.")


if __name__ == "__main__":
    main()
